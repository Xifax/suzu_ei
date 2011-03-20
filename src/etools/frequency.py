# -*- coding: utf-8 -*-
'''
Created on Mar 20, 2011

@author: Yadavito
'''

# internal #
import os, pickle

# own #
from settings.constants import *

# external #
from openpyxl.reader.excel import load_workbook

class FrequencyLookup:
    
    def __init__(self):
        self.frequency_dict = {}
        
    def loadFrequencyDict(self):
        dump = open(PATH_TO_RES + RESULTING_FREQUENCIES, 'r')
        self.frequency_dict = pickle.load(dump)
        dump.close()
        
    def parseXLSXSource(self):
         
        wb = load_workbook(filename = PATH_TO_RES + FREQ + FREQUENCIES_XL)
        ws = wb.get_active_sheet()
        
        frequency_dict = {}
        
        for i in range(1, ws.get_highest_row()):
            word = ws.cell(row = i, column = 0)
            frequency = ws.cell(row = i, column = 5)    #SUBTL(WF) frequency
            #frequency = ws.cell(row = i, column = 7)    #SUBTL(CD) frequency (0.0 ~ 100.0)
            
            if isinstance(word.value, unicode):
                if len(word.value) > 1:
                    frequency_dict[word.value] = frequency.value
            else:
                print word.value
            
            i = i + 1
        
        dump = open(PATH_TO_RES + RESULTING_FREQUENCIES, 'w')
        pickle.dump(frequency_dict, dump)
        
        del frequency_dict
        dump.close()
        
        #self.frequency_dict = frequency_dict
        
        
#        frequency_dict = {}
#
#        wb = load_workbook(filename = PATH_TO_RES + FREQ + FREQUENCIES_XL, use_iterators = True)
#        ws = wb.get_sheet_by_name(name = XL_SHEET)
#        for row in ws.iter_rows('A2:B74287'):
#            for cell in row:
#                word = cell.internal_value
#                frequency = cell.internal_value
#                
#                if isinstance(word, unicode):
#                    if len(word) > 1:
#                        frequency_dict[word] = frequency
#                else:
#                    print word
#                
#        self.frequency_dict = frequency_dict

    def parseTXTSource(self):
        pass
    
    def getFrequencyRange(self, min, max, exact = False):
        #NB: in cases values are frequencies!
        if not exact:
            return dict((keys,values) for keys, values in self.frequency_dict.iteritems() if min <= values <= max)
        else:
            return dict((keys,values) for keys, values in self.frequency_dict.iteritems() if min < values < max)
        
    def getFrequencyRangeLimits(self):
        return { 'min' : self.frequency_dict[min(self.frequency_dict, key = lambda x: self.frequency_dict.get(x) )], 
                'max': self.frequency_dict[max(self.frequency_dict, key = lambda x: self.frequency_dict.get(x) )] }
    
    def getItemsCountInRange(self, min, max):
        return len(self.getFrequencyRange(min, max))